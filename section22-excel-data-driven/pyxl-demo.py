import openpyxl

workbook = openpyxl.load_workbook("test-data.xlsx")
worksheet = workbook.active

# READ
cell_data = worksheet.cell(row=1, column=2)
print(cell_data)

# WRITE
worksheet.cell(row=2, column=2).value = "AKASH"
print(worksheet.cell(row=2, column=2).value)
print(worksheet['A5'].value)

print(worksheet.max_row)
print(worksheet.max_column)
data_dict = {}
for i in range(1, worksheet.max_row + 1):
    if worksheet.cell(row=i, column=1).value=="Test case 2":
        for j in range(2, worksheet.max_column + 1):
            print(worksheet.cell(row=i, column=j).value)
            data_dict[worksheet.cell(row=1, column=j).value] = worksheet.cell(row=i, column=j).value

print(data_dict)
